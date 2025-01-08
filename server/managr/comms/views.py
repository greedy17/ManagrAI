import json
from django.db.models.fields import FloatField
import httpx
import logging
import io
import csv
import xlrd
import pytz
from django.db.models import Q
from openpyxl import load_workbook
from rest_framework import (
    mixins,
    viewsets,
)
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from asgiref.sync import async_to_sync, sync_to_async
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from .pagination import PageNumberPagination
from datetime import datetime, timedelta, timezone
from newspaper import Article, ArticleException
from managr.api.models import ExpiringTokenAuthentication
from managr.core.models import TaskResults
from rest_framework.response import Response
from rest_framework import (
    permissions,
    mixins,
    status,
    viewsets,
)
from urllib.parse import urlencode
from django.shortcuts import redirect
from rest_framework.decorators import action
from . import constants as comms_consts
from .filters import JournalistContactFilter
from .models import (
    CompanyDetails,
    Search,
    TwitterAccount,
    Pitch,
    Process,
    InstagramAccount,
    Discovery,
    Journalist,
    EmailTracker,
    Thread,
)
from .models import Article as InternalArticle
from .models import WritingStyle, AssistAlert, JournalistContact
from managr.core.models import User
from managr.comms import exceptions as comms_exceptions
from .tasks import (
    emit_process_website_domain,
    emit_send_news_summary,
    emit_send_social_summary,
    emit_share_client_summary,
    _add_journalist_to_db,
    emit_process_contacts_excel,
    emit_process_bulk_draft,
)
from .serializers import (
    SearchSerializer,
    PitchSerializer,
    AssistAlertSerializer,
    ProcessSerializer,
    TwitterAccountSerializer,
    InstagramAccountSerializer,
    WritingStyleSerializer,
    DiscoverySerializer,
    EmailTrackerSerializer,
    JournalistContactSerializer,
    CompanyDetailsSerializer,
    ThreadSerializer,
)
from managr.core import constants as core_consts
from managr.utils.client import Variable_Client
from managr.utils.misc import decrypt_dict
from managr.core import exceptions as open_ai_exceptions
from rest_framework.decorators import (
    api_view,
    permission_classes,
    authentication_classes,
)
from managr.comms.utils import (
    generate_config,
    normalize_article_data,
    get_domain,
    extract_pdf_text,
    convert_pdf_from_url,
    extract_email_address,
    google_search,
    alternate_google_search,
    check_journalist_validity,
    get_journalists,
    merge_sort_dates,
    get_url_traffic_data,
    get_article_data,
    get_social_data,
    get_trend_articles,
    get_youtube_data,
    get_tweet_data,
    convert_social_search,
    get_bluesky_data,
)
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from managr.comms.tasks import emit_get_meta_account_info
from managr.api.emails import send_html_email, send_mailgun_email
from newspaper.article import ArticleException
from managr.slack.helpers.utils import send_to_error_channel

logger = logging.getLogger("managr")


# HELPER FUNCTIONS
def getclips(request):
    try:
        user = User.objects.get(id=request.GET.get("user_id"))
        search = request.GET.get("search", False)
        boolean = request.GET.get("boolean", False)
        date_to = request.GET.get("date_to", False)
        date_from = request.GET.get("date_from", False)
        is_report = request.GET.get("is_report", False)
        project = request.GET.get("project", None)
        suggestions = ""
        if "journalist:" in search:
            internal_articles = InternalArticle.search_by_query(search, date_to, date_from, True)
            articles = normalize_article_data([], internal_articles)
            return {"articles": articles, "string": search}
        if not boolean:
            url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
            prompt = comms_consts.OPEN_AI_QUERY_STRING(search, project)
            body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                user.email,
                prompt,
                token_amount=500,
                top_p=0.1,
            )
            with Variable_Client() as client:
                r = client.post(
                    url,
                    data=json.dumps(body),
                    headers=core_consts.OPEN_AI_HEADERS,
                )
            r = open_ai_exceptions._handle_response(r)
            query_input = r.get("choices")[0].get("message").get("content")
            news_res = Search.get_clips(query_input, date_to, date_from, is_report)
            articles = news_res["articles"]
        else:
            news_res = Search.get_clips(boolean, date_to, date_from, is_report)
            articles = news_res["articles"]
            query_input = boolean
        articles = [article for article in articles if article["title"] != "[Removed]"]
        internal_articles = InternalArticle.search_by_query(
            query_input, date_to, date_from, False, is_report
        )
        articles = normalize_article_data(articles, internal_articles, is_report)
        if not articles:
            suggestions = Search.no_results(user.email, query_input)
            query_input = suggestions.get("choices")[0].get("message").get("content")
        return {"articles": articles, "string": query_input}

    except Exception as e:
        send_to_error_channel(str(e), user.email, "get clips (platform)")
        return {"error": str(e)}


def process_journalists(journalists, contacts):
    for idx, journalist_info in enumerate(journalists):
        name_list = journalist_info["name"].split(" ")
        first = name_list[0]
        last = name_list[len(name_list) - 1]
        try:
            journalist = contacts.filter(
                journalist__first_name__icontains=first, journalist__last_name__icontains=last
            ).first()
            if journalist:
                journalists[idx]["email"] = journalist.journalist.email
            else:
                continue
        except JournalistContact.DoesNotExist as e:
            print(f"{e}: {journalist_info['name']}")
            continue
        except Exception as e:
            print(e)
            continue
    return journalists


# VIEWSETS


class PRSearchViewSet(
    viewsets.GenericViewSet,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.CreateModelMixin,
    mixins.DestroyModelMixin,
):
    authentication_classes = [ExpiringTokenAuthentication]
    serializer_class = SearchSerializer

    def get_queryset(self):
        return Search.objects.filter(user__organization=self.request.user.organization)

    def get_all_searches(self):
        return Search.objects.filter(organization=self.request.user.organization)

    def create(self, request, *args, **kwargs):
        user = request.user
        data = request.data
        data["user"] = str(user.id)
        try:
            serializer = SearchSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            response_data = serializer.data
            # serializer.instance.update_boolean()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_201_CREATED, data=response_data)

    def update(self, request, *args, **kwargs):
        search = Search.objects.get(id=request.data.get("id"))
        all_keys = [key for key, value in request.data.items()]
        keys = [key for key in all_keys]
        try:
            for field in keys:
                setattr(search, field, request.data[field])
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        search.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="summary",
    )
    def get_summary(self, request, *args, **kwargs):
        user = request.user
        if user.has_hit_summary_limit:
            return Response(status=status.HTTP_426_UPGRADE_REQUIRED)
        clips = request.data.get("clips")
        search = request.data.get("search")
        instructions = request.data.get("instructions", False)
        previous = request.data.get("previous", False)
        is_follow_up = request.data.get("followUp", False)
        trending = request.data.get("trending", False)
        company = request.data.get("company")
        if "journalist:" in search:
            instructions = comms_consts.JOURNALIST_INSTRUCTIONS(company)
        has_error = False
        attempts = 1
        token_amount = 2000
        timeout = 60.0
        while True:
            try:
                res = Search.get_summary(
                    request.user,
                    token_amount,
                    timeout,
                    clips,
                    search,
                    previous,
                    is_follow_up,
                    company,
                    trending,
                    instructions,
                    True,
                )
                message = res.get("choices")[0].get("message").get("content").replace("**", "*")
                user.add_meta_data("news_summaries")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 6000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 2000
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "news summary (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"summary": message})

        return Response(data={"summary": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="article-summary",
    )
    def get_article_summary(self, request, *args, **kwargs):
        user = request.user
        if user.has_hit_summary_limit:
            return Response(status=status.HTTP_426_UPGRADE_REQUIRED)
        url = request.data["params"]["url"]
        search = request.data["params"]["search"]
        instructions = request.data["params"]["instructions"]
        has_error = False
        attempts = 1
        token_amount = 2000
        timeout = 60.0
        while True:
            try:
                article_res = Article(url, config=generate_config())
                article_res.download()
                article_res.parse()
                text = article_res.text.replace("\n", "")
                open_ai_url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_ARTICLE_SUMMARY(
                    datetime.now().date(), text, search, instructions, True
                )
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email, prompt, model="o1-mini"
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        open_ai_url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                r = open_ai_exceptions._handle_response(r)
                message = r.get("choices")[0].get("message").get("content").replace("**", "*")
                user.add_meta_data("article_summaries")
                task = emit_process_website_domain([url], user.organization.name)
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2500:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except ArticleException:
                return Response(
                    status=status.HTTP_400_BAD_REQUEST,
                    data={"error": "We could not access that url"},
                )
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "article summary (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"summary": message})

        return Response(data={"summary": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="regenerate-article-summary",
    )
    def regenerate_article_summary(self, request, *args, **kwargs):
        user = request.user
        if user.has_hit_summary_limit:
            return Response(status=status.HTTP_426_UPGRADE_REQUIRED)
        url = request.data["params"]["url"]
        original_summary = request.data["params"]["summary"]
        instructions = request.data["params"]["instructions"]
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                article_res = Article(url, config=generate_config())
                article_res.download()
                article_res.parse()
                text = article_res.text
                open_ai_url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_REGENERATE_ARTICLE(
                    text,
                    original_summary,
                    instructions,
                )
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        open_ai_url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                r = open_ai_exceptions._handle_response(r)
                message = r.get("choices")[0].get("message").get("content").replace("**", "*")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except ArticleException:
                return Response(
                    status=status.HTTP_400_BAD_REQUEST,
                    data={"error": "We could not access that url"},
                )
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "regenerate article summary (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"summary": message})

        return Response(data={"summary": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="generate-content",
    )
    def generate_content(self, request, *args, **kwargs):
        user = request.user
        if user.has_hit_summary_limit:
            return Response(status=status.HTTP_426_UPGRADE_REQUIRED)
        url = request.data["params"]["url"]
        instructions = request.data["params"]["instructions"]
        style = request.data["params"]["style"]
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                article_res = Article(url, config=generate_config())
                article_res.download()
                article_res.parse()
                article = article_res.text

                open_ai_url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_GENERATE_CONTENT(
                    datetime.now().date(), article, "", instructions
                )
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        open_ai_url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                r = open_ai_exceptions._handle_response(r)
                message = r.get("choices")[0].get("message").get("content").replace("**", "*")

                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except ArticleException:
                return Response(
                    status=status.HTTP_400_BAD_REQUEST,
                    data={"error": "We could not access that url"},
                )
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "generate content (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"summary": message})

        return Response(data={"content": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="web-summary",
    )
    def get_web_summary(self, request, *args, **kwargs):
        user = request.user
        if user.has_hit_summary_limit:
            return Response(status=status.HTTP_426_UPGRADE_REQUIRED)
        news_url = request.data["params"]["url"]
        instructions = request.data["params"]["instructions"]
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                article_res = Article(news_url, config=generate_config())
                article_res.download()
                article_res.parse()
                text = article_res.text
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_WEB_SUMMARY(
                    datetime.now().date(),
                    text,
                    instructions,
                )
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                r = open_ai_exceptions._handle_response(r)
                message = r.get("choices")[0].get("message").get("content").replace("**", "*")
                user.add_meta_data("article_summaries")
                task = emit_process_website_domain([url], user.organization.name)
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except ArticleException:
                return Response(
                    status=status.HTTP_400_BAD_REQUEST,
                    data={"error": "We could not access that url"},
                )
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "web summary (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"summary": message})
        return Response(data={"summary": message})

    @action(
        methods=["get"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="tweets",
    )
    def get_tweets(self, request, *args, **kwargs):
        user = User.objects.get(id=request.GET.get("user_id"))
        twitter_account = user.twitter_account
        has_error = False
        search = request.GET.get("search")
        project = request.GET.get("project", None)
        query_input = None
        next_token = False
        tweet_list = []
        attempts = 1
        while True:
            try:
                if attempts >= 10:
                    break
                if query_input is None:
                    url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                    prompt = comms_consts.OPEN_AI_TWITTER_SEARCH_CONVERSION(search, project)
                    body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                        user.email,
                        prompt,
                        token_amount=500,
                        top_p=0.1,
                    )

                    with Variable_Client() as client:
                        r = client.post(
                            url,
                            data=json.dumps(body),
                            headers=core_consts.OPEN_AI_HEADERS,
                        )
                    r = open_ai_exceptions._handle_response(r)
                    query_input = r.get("choices")[0].get("message").get("content")
                    query_input = query_input.replace("AND", " ")
                    query_input = query_input + " lang:en -is:retweet"
                    if "from:" not in query_input:
                        query_input = query_input + " is:verified"
                tweet_res = twitter_account.get_tweets(query_input, next_token)

                tweets = tweet_res.get("data", None)
                includes = tweet_res.get("includes", None)
                attempts += 1
                if not tweets:
                    suggestions = twitter_account.no_results(user.email, search)
                    query_input = suggestions.get("choices")[0].get("message").get("content")
                if tweets:
                    if "next_token" in tweet_res["meta"].keys():
                        next_token = tweet_res["meta"]["next_token"]
                    user_data = tweet_res["includes"].get("users")
                    for tweet in tweets:
                        if len(tweet_list) > 39:
                            break
                        for user in user_data:
                            if user["id"] == tweet["author_id"]:
                                if user["public_metrics"]["followers_count"] > 10000:
                                    tweet["user"] = user
                                    tweet_list.append(tweet)
                                break
                if len(tweet_list) < 40 and tweets:
                    continue
                break
            except KeyError as e:
                logger.exception(e)
                return Response(
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    data={"error": f"No results for {query_input}", "string": query_input},
                )
            except comms_exceptions.TooManyRequestsError:
                if len(tweet_list):
                    break
                else:
                    return Response(
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        data={
                            "error": f"We've hit an issue contacting Twitter for {query_input}",
                            "string": query_input,
                        },
                    )
            except Exception as e:
                has_error = True
                logger.exception(e)
                tweet_res = e
                break
        if has_error:
            send_to_error_channel(tweet_res, user.email, "get tweets (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": tweet_res})
        query_input = query_input.replace("-is:retweet", "").replace("is:verified", "")
        return Response({"tweets": tweet_list, "string": query_input, "includes": includes})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="tweet-summary",
    )
    def get_tweet_summary(self, request, *args, **kwargs):
        user = request.user
        if user.has_hit_summary_limit:
            return Response(status=status.HTTP_426_UPGRADE_REQUIRED)
        tweets = request.data.get("tweets")
        search = request.data.get("search")
        company = request.data.get("company")
        instructions = request.data.get("instructions", False)
        follow_up = request.data.get("followUp", False)
        previous = request.data.get("previous", None)
        if user.has_twitter_integration:
            twitter_account = user.twitter_account
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                if follow_up:
                    if user.has_twitter_integration:
                        res = twitter_account.get_summary_follow_up(
                            request.user,
                            token_amount,
                            timeout,
                            previous,
                            tweets,
                            company,
                            instructions,
                        )
                    else:
                        res = user.get_summary_follow_up(
                            request.user,
                            token_amount,
                            timeout,
                            previous,
                            tweets,
                            company,
                            instructions,
                        )
                else:
                    if user.has_twitter_integration:
                        res = twitter_account.get_summary(
                            request.user,
                            token_amount,
                            timeout,
                            tweets,
                            search,
                            company,
                            instructions,
                            True,
                        )
                    else:
                        res = user.get_summary(
                            request.user,
                            token_amount,
                            timeout,
                            tweets,
                            search,
                            company,
                            instructions,
                            True,
                        )
                message = res.get("choices")[0].get("message").get("content").replace("**", "*")
                user.add_meta_data("tweet_summaries")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "tweet summary (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"summary": message})

        return Response(data={"summary": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="suggestions",
    )
    def get_suggestions(self, request, *args, **kwargs):
        user = request.user
        name = user.first_name
        company = user.organization
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_SEARCH_SUGGESTIONS(name, company)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)
                message = res.get("choices")[0].get("message").get("content").replace("**", "*")
                user.add_meta_data("news_summaries")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "get suggestions (platform)")
            return Response(
                status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"suggestions": message}
            )

        return Response(data={"suggestions": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="relevant-clips",
    )
    def get_most_relevant(self, request, *args, **kwargs):
        user = request.user
        social = request.data.get("social")
        term = request.data.get("term")
        clips = request.data.get("clips")
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                if social:
                    prompt = comms_consts.OPEN_AI_RELEVANT_POSTS(term, clips)
                else:
                    prompt = comms_consts.OPEN_AI_RELEVANT_ARTICLES(term, clips)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)
                message = res.get("choices")[0].get("message").get("content").replace("**", "*")
                # user.add_meta_data("most_relevant")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"data": message})

        return Response(data={"data": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="top-journalists",
    )
    def get_top_journalist(self, request, *args, **kwargs):
        user = request.user
        term = request.data.get("term")
        clips = request.data.get("clips")
        social = request.data.get("social")
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                if social:
                    prompt = comms_consts.OPEN_AI_TOP_INFLUENCERS(clips)
                else:
                    prompt = comms_consts.OPEN_AI_TOP_JOURNALISTS(term, clips)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)
                message = res.get("choices")[0].get("message").get("content").replace("**", "*")
                # user.add_meta_data("most_relevant")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "get top journalists (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"data": message})

        return Response(data={"data": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="related-topics",
    )
    def get_related_topics(self, request, *args, **kwargs):
        user = request.user
        clips = request.data.get("clips")
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_RELATED_TOPICS(clips)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)
                message = res.get("choices")[0].get("message").get("content").replace("**", "*")
                # user.add_meta_data("most_relevant")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"data": message})

        return Response(data={"data": message})

    @action(
        methods=["get"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="generate-link",
    )
    def generate_link(self, request, *args, **kwargs):
        search = Search.objects.get(id=request.GET.get("id"))
        link = search.generate_shareable_link()
        return Response(data={"link": link})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="email-summary",
    )
    def share_email_summary(self, request, *args, **kwargs):
        link = request.data.get("link", "N/A")
        email = request.data.get("email", request.user.email)
        title = request.data.get("title", "")
        emit_share_client_summary(link, title, email)
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="recipients",
    )
    def edit_recipients(self, request, *args, **kwargs):
        recipient = request.data.get("recipient")
        search_id = request.data.get("search_id")
        action = request.data.get("action")
        search = Search.objects.get(id=search_id)
        if action == "ADD":
            search.add_recipient(recipient)
        else:
            search.remove_recipient(recipient)
        return Response(status=status.HTTP_200_OK)


class PitchViewSet(
    viewsets.GenericViewSet,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
):
    serializer_class = PitchSerializer

    def get_queryset(self):
        return Pitch.objects.filter(user=self.request.user)

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            readSerializer = self.serializer_class(instance=serializer.instance)
        except Exception as e:
            logger.exception(f"Error validating data for pitch <{e}>")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_201_CREATED, data=readSerializer.data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = self.request.data
        serializer = self.serializer_class(instance=instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="generate",
    )
    def generate_pitch(self, request, *args, **kwargs):
        user = request.user
        if user.has_hit_summary_limit:
            return Response(status=status.HTTP_426_UPGRADE_REQUIRED)
        type = request.data.get("type")
        instructions = request.data.get("instructions")
        style = request.data.get("style")
        pitch_id = request.data.get("pitch_id", False)

        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0

        while True:
            try:
                res = Pitch.generate_pitch(user, type, instructions, style, token_amount, timeout)
                pitch = res.get("choices")[0].get("message").get("content")
                if pitch_id:
                    saved_pitch = Pitch.objects.get(id=pitch_id)
                    saved_pitch.generate_pitch = pitch
                    saved_pitch.save()
                user.add_meta_data("pitches")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount >= 3000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 1000
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "generate pitch (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response({"pitch": pitch})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="regenerate",
    )
    def get_regenerated_pitch(self, request, *args, **kwargs):
        user = request.user
        instructions = request.data.get("instructions")
        pitch = request.data.get("pitch")
        style = request.data.get("style")
        details = request.data.get("details")
        elma = core_consts.ELMA
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0

        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_PTICH_DRAFT_WITH_INSTRUCTIONS(
                    elma, pitch, instructions, style, details
                )
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                r = open_ai_exceptions._handle_response(r)
                pitch = r.get("choices")[0].get("message").get("content")
                user.add_meta_data("pitches")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "regenerate pitch (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response({"pitch": pitch})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="learn",
    )
    def learn_writing_style(self, request, *args, **kwargs):
        user = request.user
        example = request.data["params"]["example"]
        title = request.data["params"]["title"]
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0

        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_LEARN_WRITING_STYLE_PROMPT(example)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                r = open_ai_exceptions._handle_response(r)
                style = r.get("choices")[0].get("message").get("content")
                writing_dict = {"title": title, "style": style, "user": request.user}
                WritingStyle.objects.create(**writing_dict)
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response({"style": style})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="delete-style",
    )
    def delete_writing_style(self, request, *args, **kwargs):
        style_id = request.data["params"]["style_id"]
        style = WritingStyle.objects.get(id=style_id)
        style.delete()
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="journalists",
    )
    def find_journalist(self, request, *args, **kwargs):
        user = request.user
        # if user.has_hit_summary_limit:
        #     return Response(status=status.HTTP_426_UPGRADE_REQUIRED)
        type = request.data.get("type")
        beat = request.data.get("beat")
        location = request.data.get("location")
        content = request.data.get("content")
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0

        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_FIND_JOURNALISTS(type, beat, location, content)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)

                journalists = res.get("choices")[0].get("message").get("content")

                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "find journalists (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response({"journalists": journalists})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="feedback",
    )
    def get_feedback(self, request, *args, **kwargs):
        user = request.user
        type = request.data.get("type")
        audience = request.data.get("audience")
        objective = request.data.get("objective")
        feedback = request.data.get("feedback")
        content = request.data.get("content")
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0

        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_GENERATE_FEEDBACK(
                    type, audience, objective, feedback, content
                )
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)

                feedback = res.get("choices")[0].get("message").get("content")

                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "get feedback (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response({"feedback": feedback})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="regenerate-feedback",
    )
    def regenerate_with_feedback(self, request, *args, **kwargs):
        user = request.user
        feedback = request.data.get("feedback")
        content = request.data.get("content")
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0

        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.REGENERATE_CONTENT_WITH_FEEDBACK(content, feedback)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)

                feedback = res.get("choices")[0].get("message").get("content")

                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response({"feedback": feedback})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="rewrite",
    )
    def rewrite_pitch(self, request, *args, **kwargs):
        user = request.user
        original = request.data.get("original")
        bio = request.data.get("bio")
        style = request.data.get("style", None)
        with_style = request.data.get("with_style", False)
        journalist = request.data.get("journalist", None)
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0

        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_REWRITE_PTICH(
                    original, bio, style, with_style, journalist, user.first_name
                )
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    model="o1-mini",
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                r = open_ai_exceptions._handle_response(r)
                pitch = (
                    r.get("choices")[0]
                    .get("message")
                    .get("content")
                    .replace("json", "")
                    .replace("```", "")
                )
                pitch = json.loads(pitch)
                user.add_meta_data("emailDraft")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response(data=pitch)


class AssistAlertViewSet(
    viewsets.GenericViewSet,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
):
    serializer_class = AssistAlertSerializer

    def get_queryset(self):
        return AssistAlert.objects.filter(user=self.request.user)

    def create(self, request, *args, **kwargs):
        datetime = request.data.pop("run_at")
        request.data["run_at"] = datetime
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            readSerializer = self.serializer_class(instance=serializer.instance)
        except Exception as e:
            logger.exception(f"Error validating data for email alert <{e}>")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_201_CREATED, data=readSerializer.data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = self.request.data
        serializer = self.serializer_class(instance=instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="test-alert",
    )
    def test_email_alert(self, request, *args, **kwargs):
        alert_id = request.data.get("alert_id")
        social = request.data.get("social")
        if social:
            emit_send_social_summary(alert_id, str(datetime.now()))
        else:
            emit_send_news_summary(alert_id, str(datetime.now()))
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["get"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="get-email-alerts",
    )
    def get_email_alerts(self, request, *args, **kwargs):
        alerts = AssistAlert.objects
        serialized = AssistAlertSerializer(alerts, many=True)
        return Response(data=serialized.data, status=status.HTTP_200_OK)


class ProcessViewSet(
    viewsets.GenericViewSet,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
):
    serializer_class = ProcessSerializer

    def get_queryset(self):
        return Process.objects.filter(user=self.request.user)

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            readSerializer = self.serializer_class(instance=serializer.instance)
        except Exception as e:
            logger.exception(f"Error validating data for detials <{e}>")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_201_CREATED, data=readSerializer.data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = self.request.data
        serializer = self.serializer_class(instance=instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="run",
    )
    def run_process(self, request, *args, **kwargs):
        user = request.user
        type = request.data.get("type")
        summary = request.data.get("summary")
        details = request.data.get("details")
        process_id = request.data.get("process_id")
        style = request.data.get("style")

        #         style = """
        #         The author's style is formal yet personal, employing a tone of gratitude and excitement. They use first-person narrative, making the text relatable and engaging. The structure is chronological, starting with the present and moving to past achievements, ending with a hopeful look to the future. The author uses sophisticated vocabulary and complex sentence structures, demonstrating a high level of education and professionalism.
        # The author establishes credibility by mentioning specific names, roles, and institutions, and by expressing gratitude towards mentors and colleagues. They avoid promotional language, focusing instead on personal growth and learning opportunities.
        # Guidelines for replicating this style:
        # 1. Use a formal yet personal tone, employing first-person narrative.
        # 2. Structure the text chronologically, linking past, present, and future.
        # 3. Use sophisticated vocabulary and complex sentence structures.
        # 4. Establish credibility by mentioning specific names, roles, and institutions.
        # 5. Express gratitude and excitement about learning and growth opportunities.
        # 6. Avoid promotional language, focusing on informative and trustworthy discourse.
        #         """

        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0

        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.RUN_PROCESS(datetime.now(), type, summary, details, style)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)

                content = res.get("choices")[0].get("message").get("content")

                user.add_meta_data("assist")

                if process_id:
                    process = Process.objects.get(id=process_id)
                    process.generated_content = content
                    process.save()
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response({"content": content})


class DiscoveryViewSet(
    viewsets.GenericViewSet,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
):
    serializer_class = DiscoverySerializer

    def get_queryset(self):
        return Discovery.objects.filter(user__organization=self.request.user.organization)

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            readSerializer = self.serializer_class(instance=serializer.instance)
        except Exception as e:
            logger.exception(f"Error validating data for details <{e}>")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_201_CREATED, data=readSerializer.data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = self.request.data
        serializer = self.serializer_class(instance=instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="run",
    )
    def run_discovery(self, request, *args, **kwargs):
        user = request.user
        info = request.data.get("info")
        content = request.data.get("content", None)
        previous = request.data.get("previous", None)
        message = get_journalists(info, content, previous)
        if isinstance(message, str):
            send_to_error_channel(message, user.email, "run discovery (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=message)
        user.add_meta_data("discovery")
        return Response(data={"journalists": message})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="email",
    )
    def send_email(self, request, *args, **kwargs):
        user = request.user
        subject = request.data.get("subject")
        body = request.data.get("body").replace("[Your Name]", f"\n\n{user.full_name}")
        recipient = request.data.get("recipient")
        name = request.data.get("name")
        cc = request.data.get("cc", [])
        bcc = request.data.get("bcc", [])
        if cc:
            cc = [cc]
        if bcc:
            bcc = [bcc]
        draftId = request.data.get("draftId", None)
        if draftId:
            tracker = EmailTracker.objects.filter(id=draftId).first()
            if tracker:
                user = tracker.user
        if user.has_google_integration or user.has_microsoft_integration:
            res = user.email_account.send_email(recipient, subject, body, name, cc, bcc, draftId)
            user.add_meta_data("emailSent")
        else:
            res = send_mailgun_email(user, name, subject, recipient, body, bcc, cc, draftId)
        sent = res["sent"]
        if sent:
            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            return Response(
                status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": res["error"]}
            )

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="verify",
    )
    def verify_email(self, request, *args, **kwargs):
        user = request.user
        journalist = request.data.get("journalist").strip()
        outlet = request.data.get("publication")
        email = request.data.get("email")
        name_list = journalist.strip().split(" ")
        first = name_list[0]
        last = name_list[len(name_list) - 1]
        try:
            email_check = Journalist.objects.filter(
                Q(email=email) | Q(first_name=first, last_name=last)
            )
            if email_check.first():
                internal_journalist = email_check.first()
                if not internal_journalist.date_verified:
                    score = Journalist.verify_email(internal_journalist.email)
                    is_valid = True if score >= 85 else False
                    internal_journalist.score = score
                    internal_journalist.verified = is_valid
                    internal_journalist.date_verified = datetime.now()
                    internal_journalist.save()
                return Response(
                    status=status.HTTP_200_OK,
                    data={
                        "is_valid": internal_journalist.verified,
                        "email": internal_journalist.email,
                    },
                )
            else:
                score = Journalist.verify_email(email)
                is_valid = True if score >= 85 else False
                if is_valid is False:
                    r = Journalist.email_finder(first, last, outlet=outlet)
                    score = r["score"]
                    if score is None:
                        score = 0
                    is_valid = True if score >= 85 else False
                    if r["email"] is not None:
                        email = r["email"]
                        request.data["email"] = email
                        request.data["publication"] = r["company"]
                request.data["accuracy_score"] = score
                user.add_meta_data("verify")
                _add_journalist_to_db(request.data, is_valid)
                return Response(
                    status=status.HTTP_200_OK, data={"is_valid": is_valid, "email": email}
                )
        except Exception as e:
            logger.exception(str(e))
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="web-context",
    )
    def journalist_web_context(self, request, *args, **kwargs):
        user = request.user
        journalist = request.data.get("journalist")
        outlet = request.data.get("outlet")
        company = request.data.get("company")
        search = request.data.get("search")
        social = request.data.get("social")
        if social:
            query = journalist
        else:
            query = f"Journalist AND {journalist} AND {outlet}"
        google_results = google_search(query)
        if len(google_results) == 0:
            return Response(
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                data={"error": "No results could be found."},
            )
        results = google_results["results"]
        images = google_results["images"]
        art = Article(results[0]["link"], config=generate_config())
        try:
            art.download()
            art.parse()
            text = art.text
        except ArticleException:
            text = ""
        except Exception:
            text = ""
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                if social:
                    prompt = comms_consts.OPEN_AI_SOCIAL_BIO(journalist, company, results, text)
                else:
                    prompt = comms_consts.OPEN_AI_RESULTS_PROMPT(journalist, results, company, text)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                    response_format={"type": "json_object"},
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)
                message = res.get("choices")[0].get("message").get("content")
                message = json.loads(message)
                message["images"] = images
                user.add_meta_data("bio")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "journalist web context (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=message)

        return Response(message)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="draft",
    )
    def draft_pitch(self, request, *args, **kwargs):
        user = request.user
        username = request.data.get("user")
        org = request.data.get("org")
        bio = request.data.get("bio")
        style = request.data.get("style")
        author = request.data.get("author")
        outlet = request.data.get("outlet")
        headline = request.data.get("headline")
        description = request.data.get("description")
        date = request.data.get("date")
        has_error = False
        attempts = 1
        token_amount = 1000
        timeout = 60.0
        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_EMAIL_JOURNALIST(
                    username, org, style, bio, author, outlet, headline, description, date
                )
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)

                message = res.get("choices")[0].get("message").get("content").replace("**", "*")
                user.add_meta_data("emailDraft")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount <= 2000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 500
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            send_to_error_channel(message, user.email, "draft pitch (platform)")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=message)

        return Response(data=message)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="network",
    )
    def network_discover(self, request, *args, **kwargs):
        user = request.user
        pitch = request.data.get("pitch", False)
        has_error = False
        attempts = 1
        token_amount = 2000
        timeout = 60.0
        while True:
            try:
                contacts = JournalistContact.objects.filter(user=user)
                contact_list = [
                    f"Name:{contact.journalist.full_name}, Outlet:{contact.journalist.outlet}"
                    for contact in contacts
                ]
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                prompt = comms_consts.OPEN_AI_PITCH_JOURNALIST_LIST(contact_list, pitch)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    "You are a VP of Communications",
                    token_amount=token_amount,
                    top_p=0.1,
                    response_format={"type": "json_object"},
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)
                res_content = json.loads(res.get("choices")[0].get("message").get("content"))
                journalists_res = res_content.get("journalists")
                journalists = process_journalists(journalists_res, contacts)
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount >= 5000:
                    has_error = True
                    message = "Token amount error"
                    break
                else:
                    token_amount += 1000
                    continue
            except httpx.ReadTimeout as e:
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=message)

        return Response(data={"journalists": journalists})


class JournalistContactViewSet(
    viewsets.GenericViewSet,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.CreateModelMixin,
    mixins.DestroyModelMixin,
):
    authentication_classes = [ExpiringTokenAuthentication]
    serializer_class = JournalistContactSerializer
    pagination_class = PageNumberPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = JournalistContactFilter
    ordering_fields = ["journalist__email", "journalist__first_name", "journalist__last_name"]

    def get_queryset(self):
        # contacts = JournalistContact.objects.for_user(user=self.request.user)
        user = self.request.user
        contacts = JournalistContact.objects.filter(user__organization=user.organization).order_by(
            "-datetime_created"
        )
        search_term = self.request.query_params.get("search", "")
        tags = self.request.query_params.getlist("tags[]", [])
        user_id = self.request.query_params.get("user_id", "")

        if user_id:
            contacts = contacts.filter(user=user_id)

        if tags:
            tag_queries = Q()
            for tag in tags:
                tag_queries |= Q(tags__contains=[tag])
            contacts = contacts.filter(tag_queries).distinct()

        if search_term:
            contacts = contacts.filter(
                Q(journalist__email__icontains=search_term)
                | Q(journalist__first_name__icontains=search_term)
                | Q(journalist__last_name__icontains=search_term)
                | Q(journalist__outlet__icontains=search_term)
            )
        return contacts

    def create(self, request, *args, **kwargs):
        user = request.user
        journalist = request.data.pop("journalist").strip()
        email = request.data.pop("email").strip()
        outlet = request.data.pop("outlet").strip()
        journalist = check_journalist_validity(journalist, outlet, email)
        if isinstance(journalist, dict) and "error" in journalist.keys():
            return Response(
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                data={"error": "Could not create contact"},
            )
        else:
            request.data["journalist"] = journalist.id
            request.data["user"] = request.user.id
            try:
                serializer = self.serializer_class(data=request.data)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                readSerializer = self.serializer_class(instance=serializer.instance)
                # user.add_meta_data("contacts")
            except Exception as e:
                send_to_error_channel(
                    str(e), user.email, f"creating journalist contact (platform):\n{request.data}"
                )
                logger.exception(f"Error validating data for details <{e}>")
                return Response(
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)}
                )
            return Response(status=status.HTTP_201_CREATED, data=readSerializer.data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.serializer_class(instance, data=request.data, partial=True)
        if serializer.is_valid(raise_exception=True):
            serializer.update(instance=instance, validated_data=request.data)
        else:
            return Response(
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                data={"error": "Error updating contact"},
            )

        return Response(status=status.HTTP_200_OK, data=serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="modify_tags",
    )
    def modify_tags(self, request, *args, **kwargs):
        ids = request.data.get("ids")
        tag = request.data.get("tag")
        modifier = request.data.get("modifier")
        for id in ids:
            JournalistContact.modify_tags(id, tag, modifier)
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="add-note",
    )
    def add_note(self, request, *args, **kwargs):
        contact_id = request.data.get("id")
        note = request.data.get("note")
        contact = JournalistContact.objects.get(id=contact_id)
        if contact.notes is None:
            contact.notes = []
        timezone = pytz.timezone(self.request.user.timezone)
        date = datetime.now(pytz.utc)
        current_time = date.astimezone(timezone).isoformat()
        note_data = {"date": current_time, "note": note, "user": request.user.full_name}
        contact.notes.append(note_data)
        contact.save()
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="edit-note",
    )
    def edit_note(self, request, *args, **kwargs):
        contact_id = request.data.get("id")
        note_index = request.data.get("note_index")
        new_note = request.data.get("note")

        contact = JournalistContact.objects.get(id=contact_id)

        timezone = pytz.timezone(self.request.user.timezone)
        date = datetime.now(pytz.utc)
        current_time = date.astimezone(timezone).isoformat()

        contact.notes[note_index]["note"] = new_note
        contact.notes[note_index]["date_modified"] = current_time
        contact.notes[note_index]["modified_by"] = request.user.full_name

        contact.save()
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="delete-note",
    )
    def delete_note(self, request, *args, **kwargs):
        contact_id = request.data.get("id")
        note_index = request.data.get("note_index")
        contact = JournalistContact.objects.get(id=contact_id)
        contact.notes.pop(note_index)
        contact.save()
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["get"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="tag_list",
    )
    def get_tag_list(self, request, *args, **kwargs):
        user = request.user
        tags = JournalistContact.get_tags_by_user(user)
        return Response(status=status.HTTP_200_OK, data={"tags": tags})

    @action(
        methods=["get"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="activity",
    )
    def get_activity(self, request, *args, **kwargs):
        user = request.user
        journalist_email = request.GET.get("email")
        contacts = JournalistContact.objects.filter(
            user__organization=user.organization, journalist__email=journalist_email
        )
        trackers = EmailTracker.objects.filter(
            user__organization=user.organization, recipient=journalist_email
        )
        activity_list = []
        for contact in contacts:
            if contact.notes:
                activity_list.extend(contact.notes)
        for tracker in trackers:
            tracker_list = []
            activity = tracker.activity_log
            for event in activity:
                evt, date = event.split("|")
                event_data = {"date": date, "event": evt}
                if evt == "sent":
                    event_data["user"] = tracker.user.full_name
                tracker_list.append(event_data)
            activity_list.extend(tracker_list)
        sorted_activity_list = merge_sort_dates(activity_list, "date")
        return Response(status=status.HTTP_200_OK, data={"activity": sorted_activity_list})

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="edit",
    )
    def edit_contact(self, request, *args, **kwargs):
        email = request.data.get("email", None)
        outlet = request.data.get("outlet", None)
        contact_id = request.data.get("id")
        contact = JournalistContact.objects.get(id=contact_id)
        journalist = contact.journalist
        edit_details = f"Changes by {request.user.email}:"
        if email:
            contact.email = email
            edit_details += f" {email}"
        if outlet:
            contact.outlet = outlet
            edit_details += f" {outlet}"
        contact.save()
        journalist.needs_review = True
        if journalist.review_details:
            journalist.review_details += edit_details
        else:
            journalist.review_details = edit_details
        journalist.save()
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="insight",
    )
    def get_insight(self, request, *args, **kwargs):
        notes = request.data.get("notes", None)
        activity = request.data.get("activity", None)
        bio = request.data.get("bio", None)
        instructions = request.data.get("instructions", None)
        tracker = request.data.get("is_tracker", False)
        user = self.request.user

        has_error = False
        attempts = 1
        token_amount = 4000
        timeout = 60.0
        while True:
            try:
                url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
                if tracker:
                    prompt = comms_consts.OPEN_AI_TRACKER_INSIGHTS(notes, instructions)
                else:
                    prompt = comms_consts.OPEN_AI_GET_INSIGHTS(notes, activity, bio, instructions)
                body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                    user.email,
                    prompt,
                    token_amount=token_amount,
                    top_p=0.1,
                )
                with Variable_Client(timeout) as client:
                    r = client.post(
                        url,
                        data=json.dumps(body),
                        headers=core_consts.OPEN_AI_HEADERS,
                    )
                res = open_ai_exceptions._handle_response(r)

                content = res.get("choices")[0].get("message").get("content")
                break
            except open_ai_exceptions.StopReasonLength:
                logger.exception(
                    f"Retrying again due to token amount, amount currently at: {token_amount}"
                )
                if token_amount >= 8000:
                    has_error = True

                    message = "Token amount error"
                    break
                else:
                    token_amount += 1000
                    continue
            except httpx.ReadTimeout as e:
                print(e)
                timeout += 30.0
                if timeout >= 120.0:
                    has_error = True
                    message = "Read timeout issue"
                    logger.exception(f"Read timeout from Open AI {e}")
                    break
                else:
                    attempts += 1
                    continue
            except Exception as e:
                has_error = True
                message = f"Unknown exception: {e}"
                logger.exception(e)
                break
        if has_error:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
        return Response({"content": content})


class CompanyDetailsViewSet(
    viewsets.GenericViewSet,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.CreateModelMixin,
    mixins.DestroyModelMixin,
):
    serializer_class = CompanyDetailsSerializer

    def get_queryset(self):
        details = CompanyDetails.objects.filter(user=self.request.user)
        return details

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            readSerializer = self.serializer_class(instance=serializer.instance)
        except Exception as e:
            logger.exception(f"Error validating data for details <{e}>")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_201_CREATED, data=readSerializer.data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = self.request.data
        serializer = self.serializer_class(instance=instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="delete-details",
    )
    def delete_details(self, request, *args, **kwargs):
        detail_id = request.data["params"]["id"]
        detail = CompanyDetails.objects.get(id=detail_id)
        detail.delete()
        return Response(status=status.HTTP_200_OK)


class EmailTrackerViewSet(
    viewsets.GenericViewSet,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.CreateModelMixin,
    mixins.DestroyModelMixin,
):
    authentication_classes = [ExpiringTokenAuthentication]
    serializer_class = EmailTrackerSerializer

    def get_queryset(self):
        user = self.request.user
        org = user.organization
        trackers = EmailTracker.objects.filter(user__organization=org).order_by("-datetime_created")
        return trackers

    def create(self, request, *args, **kwargs):
        user = request.user
        try:
            serializer = EmailTrackerSerializer(
                data={
                    "user": user.id,
                    "recipient": request.data.get("recipient"),
                    "body": request.data.get("body"),
                    "subject": request.data.get("subject"),
                    "name": request.data.get("name"),
                }
            )
            serializer.is_valid(raise_exception=True)
            serializer.save()
            serializer.instance.add_activity("draft_created")
            return Response(status=status.HTTP_201_CREATED, data={"tracker": serializer.data})
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = self.request.data
        serializer = self.serializer_class(instance=instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        detail=False,
        url_path="bulk-draft",
    )
    def draft_bulk_emails(self, request, *args, **kwargs):
        result = TaskResults.objects.create(
            function_name="emit_process_bulk_draft", user_id=str(request.user.id)
        )
        task = emit_process_bulk_draft(request.data, str(request.user.id), str(result.id))
        result.task = task
        result.save()
        data = {"task_id": str(result.id)}
        return Response(status=status.HTTP_200_OK, data=data)


class ThreadViewSet(
    viewsets.GenericViewSet,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
):
    serializer_class = ThreadSerializer

    def get_queryset(self):
        return Thread.objects.filter(user=self.request.user)

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
        except Exception as e:
            logger.exception(f"Error validating data for new thread <{e}>")
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_201_CREATED, data=serializer.data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = self.request.data
        serializer = self.serializer_class(instance=instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
        except Exception as e:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": str(e)})
        return Response(status=status.HTTP_200_OK)

    @action(
        methods=["get"],
        permission_classes=[permissions.AllowAny],
        detail=False,
        url_path="shared",
    )
    def get_shared_thread(self, request, *args, **kwargs):
        user = request.user
        encrypted_code = request.GET.get("code")
        # encrypted_code = base64.urlsafe_b64decode(encrypted_code.encode('utf-8'))
        try:
            decrypted_dict = decrypt_dict(encrypted_code)
            id = decrypted_dict.get("id")
            date = decrypted_dict.get("created_at")
            report = Thread.objects.get(id=id)
            serializer = self.get_serializer(report)
            # user.add_meta_data("shared_thread")
        except Exception as e:
            print("exception is here --- >", e)
            return Response(
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                data={"error": "There was an issue retrieving this thread"},
            )
        return Response(status=status.HTTP_200_OK, data={"data": serializer.data, "date": date})


# ENDPOINTS


@require_http_methods(["GET"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
@async_to_sync
async def get_clips(request, *args, **kwargs):
    response = await sync_to_async(getclips)(request)
    return JsonResponse(data=response)


@api_view(["GET"])
@permission_classes(
    [
        permissions.AllowAny,
    ]
)
def get_shared_summary(request, encrypted_param):
    decrypted_dict = decrypt_dict(encrypted_param)
    created_at = datetime.strptime(decrypted_dict.get("created_at"), "%Y-%m-%d %H:%M:%S.%f")
    time_difference = datetime.now() - created_at
    twenty_four_hours = timedelta(hours=24)
    if time_difference > twenty_four_hours:
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    search = Search.objects.get(id=decrypt_dict["id"])
    return Response(data={"summary": search.summary})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def upload_link(request):
    url = request.data["params"]["url"]
    try:
        article_res = Article(url, config=generate_config())
        article_res.download()
        article_res.parse()

        title = article_res.title
        author = article_res.authors
        image = article_res.top_image
        date = article_res.publish_date
        text = article_res.meta_description
        domain = get_domain(url)
        article = {}
        article = {
            "title": title,
            "source": domain,
            "author": author,
            "image_url": image,
            "publish_date": date,
            "description": text,
            "link": url,
        }
        emit_process_website_domain([url], request.user.organization.name)
    except Exception as e:
        logger.exception(e)
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    return Response(data=article)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_twitter_request_token(request):
    res = TwitterAccount.get_token(request)
    return Response(res)


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_twitter_auth_link(request):
    link = TwitterAccount.get_authorization(request.token)
    return Response({"link": link})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_twitter_authentication(request):
    user = request.user
    data = request.data
    access_token = TwitterAccount.authenticate(data.get("oauth_token"), data.get("oauth_verifier"))
    data = {
        "user": user.id,
        "access_token": access_token.get("oauth_token"),
        "access_token_secret": access_token.get("oauth_token_secret"),
    }
    existing = TwitterAccount.objects.filter(user=request.user).first()
    if existing:
        serializer = TwitterAccountSerializer(data=data, instance=existing)
    else:
        serializer = TwitterAccountSerializer(data=data)
    try:
        serializer.is_valid(raise_exception=True)
        serializer.save()

    except Exception as e:
        logger.exception(str(e))
        return Response(data={"success": False})
    return Response(data={"success": True})


@api_view(["DELETE"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def revoke_twitter_auth(request):
    user = request.user
    twitter_account = TwitterAccount.objects.filter(user=user)
    try:
        twitter_account.delete()
    except Exception as e:
        return Response({"error": str(e)})
    return Response(data={"success": True})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_instagram_request_token(request):
    res = InstagramAccount.get_token(request)
    return Response(res)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_instagram_authentication(request):
    user = request.user
    data = request.data
    access_token_response = InstagramAccount.authenticate(data.get("code"))
    access_token_response["user"] = user.id
    existing = InstagramAccount.objects.filter(user=request.user).first()
    if existing:
        serializer = InstagramAccountSerializer(data=access_token_response, instance=existing)
    else:
        serializer = InstagramAccountSerializer(data=access_token_response)
    try:
        serializer.is_valid(raise_exception=True)
        serializer.save()
        emit_get_meta_account_info(str(user.id))
    except Exception as e:
        logger.exception(str(e))
        return Response(data={"success": False})
    return Response(data={"success": True})


@api_view(["DELETE"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def revoke_instagram_auth(request):
    user = request.user
    ig_account = InstagramAccount.objects.filter(user=user)
    try:
        ig_account.delete()
    except Exception as e:
        return Response({"error": str(e)})
    return Response(data={"success": True})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_email_request_token(request):
    res = InstagramAccount.get_token(request)
    return Response(res)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_email_authentication(request):
    user = request.user
    data = request.data
    access_token_response = InstagramAccount.authenticate(data.get("code"))
    access_token_response["user"] = user.id
    existing = InstagramAccount.objects.filter(user=request.user).first()
    if existing:
        serializer = InstagramAccountSerializer(data=access_token_response, instance=existing)
    else:
        serializer = InstagramAccountSerializer(data=access_token_response)
    try:
        serializer.is_valid(raise_exception=True)
        serializer.save()
        emit_get_meta_account_info(str(user.id))
    except Exception as e:
        logger.exception(str(e))
        return Response(data={"success": False})
    return Response(data={"success": True})


@api_view(["DELETE"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def revoke_email_auth(request):
    user = request.user
    ig_account = InstagramAccount.objects.filter(user=user)
    try:
        ig_account.delete()
    except Exception as e:
        return Response({"error": str(e)})
    return Response(data={"success": True})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def upload_pdf(request):
    user = request.user
    url = request.data.get("url", None)
    pdf_file = request.FILES.get("pdf_file", None)
    instructions = request.data.get("instructions", "Create a summary")
    if pdf_file:
        text, images = extract_pdf_text(pdf_file)
    else:
        text, images = convert_pdf_from_url(url)
    if not len(text):
        return Response(
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            data={"error": "No text could be extract from the uploaded PDF"},
        )
    has_error = False
    attempts = 1
    token_amount = 4000
    timeout = 60.0
    while True:
        try:
            url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
            prompt = comms_consts.OPEN_AI_GENERATE_PDF_SUMMARY(
                datetime.now().date(), instructions, text
            )
            body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                user.email,
                prompt,
                token_amount=token_amount,
                top_p=0.1,
            )
            with Variable_Client(timeout) as client:
                r = client.post(
                    url,
                    data=json.dumps(body),
                    headers=core_consts.OPEN_AI_HEADERS,
                )
            res = open_ai_exceptions._handle_response(r)

            content = res.get("choices")[0].get("message").get("content")
            break
        except open_ai_exceptions.StopReasonLength:
            logger.exception(
                f"Retrying again due to token amount, amount currently at: {token_amount}"
            )
            if token_amount <= 8000:
                has_error = True

                message = "Token amount error"
                break
            else:
                token_amount += 1000
                continue
        except httpx.ReadTimeout as e:
            print(e)
            timeout += 30.0
            if timeout >= 120.0:
                has_error = True
                message = "Read timeout issue"
                logger.exception(f"Read timeout from Open AI {e}")
                break
            else:
                attempts += 1
                continue
        except Exception as e:
            has_error = True
            message = f"Unknown exception: {e}"
            logger.exception(e)
            break
    if has_error:
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
    return Response({"content": content})


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_writing_styles(request):
    writing_styles = WritingStyle.objects.filter(user__organization=request.user.organization)
    serializer = WritingStyleSerializer(writing_styles, many=True)  # Serialize the queryset

    return Response(serializer.data)


@api_view(["GET"])
@permission_classes([])
def email_tracking_endpoint(request):
    event_type = request.GET.get("type")
    message_id = request.GET.get("id")
    try:
        tracker = EmailTracker.objects.get(id=message_id)
        if event_type == "opened":
            response = HttpResponse(content_type="image/png")
            response.write(
                b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
                b"\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDAT\x08\xd7c`\x00\x00"
                b"\x00\x02\x00\x01\xe2!\xbc\x33\x00\x00\x00\x00IEND\xaeB`\x82"
            )
            response["Content-Length"] = len(response.content)
            last_log = tracker.activity_log[len(tracker.activity_log) - 1]
            event, time = last_log.split("|")
            if event == "opened":
                message_timestamp = datetime.now().timestamp()
                datetime_obj = datetime.strptime(time, "%Y-%m-%dT%H:%M:%S.%f%z")
                unix_time = datetime_obj.timestamp()
                if unix_time - message_timestamp < 60:
                    raise Exception()
                else:
                    tracker.opens += 1
            else:
                tracker.opens += 1
        elif event_type == "delivered":
            tracker.received = True
        elif event_type == "failed":
            tracker.failed = True
        elif event_type == "clicked":
            original_url = request.GET.get("redirect")
            response = original_url
            tracker.clicks += 1
        tracker.save()
        tracker.add_activity(event_type)
    except EmailTracker.DoesNotExist:
        pass
    except Exception as e:
        logger.exception(f"{e}, {message_id}")
    if event_type == "opened":
        return response
    else:
        if "https" not in response:
            response = "https://" + response
        return HttpResponseRedirect(response)


@api_view(["POST"])
@permission_classes([])
def mailgun_webhooks(request):
    event_data = request.data["event-data"]
    message_id = event_data["message"]["headers"]["message-id"]
    event_type = event_data["event"]
    try:
        tracker = EmailTracker.objects.get(message_id=message_id)
        if event_type == "opened":
            last_log = tracker.activity_log[len(tracker.activity_log) - 1]
            event, time = last_log.split("|")
            if event == "opened":
                message_timestamp = event_data["timestamp"]
                datetime_obj = datetime.strptime(time, "%Y-%m-%dT%H:%M:%S.%f")
                unix_time = datetime_obj.timestamp()
                if unix_time - message_timestamp < 60:
                    Response(status=status.HTTP_202_ACCEPTED)
                else:
                    tracker.opens += 1
            else:
                tracker.opens += 1
        elif event_type == "delivered":
            tracker.received = True
        elif event_type == "failed":
            tracker.failed = True
        elif event_type == "clicked":
            tracker.clicks += 1
        tracker.save()
        tracker.add_activity(event_type)
    except EmailTracker.DoesNotExist:
        return Response(status=status.HTTP_202_ACCEPTED)
    except Exception as e:
        logger.exception(f"{e}, {message_id}\n{event_data}")
    return Response(status=status.HTTP_202_ACCEPTED)


@api_view(["POST"])
@permission_classes([])
def email_recieved_webhook(request):
    subject = request.POST.get("Subject")
    email_html = request.POST.get("stripped-html")
    to_email = request.POST.get("To")
    to_email = extract_email_address(to_email)
    from_email = request.POST.get("From")
    from_email = extract_email_address(from_email)
    name, domain = to_email.split("@")
    first, last = name.split(".")
    original_subject = subject.replace("Re: ", "")
    user = User.objects.get(first_name=first, last_name=last)
    try:
        trackers = EmailTracker.objects.filter(recipient__icontains=from_email, user=user)
        filtered_trackers = [email for email in trackers if email.subject in original_subject]
        if len(filtered_trackers):
            tracker = filtered_trackers[0]
            tracker.replies += 1
            tracker.recieved = True
            tracker.save()
            tracker.add_activity("reply")
    except Exception as e:
        print(e)
    email = send_html_email(
        subject,
        "core/email-templates/reply-email.html",
        from_email,
        user.email,
        {"html": email_html},
    )
    return Response(status=status.HTTP_202_ACCEPTED)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_google_summary(request):
    user = request.user
    query = request.data.get("query")
    instructions = request.data.get("instructions")
    summary = request.data.get("summary", None)
    results = request.data.get("results", None)
    project = request.data.get("project", None)
    text = ""
    elma = core_consts.ELMA
    if not instructions or not summary:
        res = alternate_google_search(query)
        if len(res) == 0:
            return Response(
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                data={"message": "No results could be found."},
            )
        results = res["results"]
        art = Article(results[0]["link"], config=generate_config())
        try:
            art.download()
            art.parse()
            text = art.text
        except ArticleException:
            text = ""
        except Exception:
            text = ""
    has_error = False
    attempts = 1
    token_amount = 1000
    timeout = 60.0
    while True:
        try:
            url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
            prompt = comms_consts.OPEN_AI_WEB_SUMMARY(
                query, results, text, instructions, summary, elma, project
            )
            body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                user.email,
                prompt,
                "You are a VP of Communications",
                token_amount=token_amount,
                top_p=0.1,
            )
            with Variable_Client(timeout) as client:
                r = client.post(
                    url,
                    data=json.dumps(body),
                    headers=core_consts.OPEN_AI_HEADERS,
                )
            res = open_ai_exceptions._handle_response(r)

            message = res.get("choices")[0].get("message").get("content").replace("**", "*")
            user.add_meta_data("google_search")
            break
        except open_ai_exceptions.StopReasonLength:
            if token_amount <= 2000:
                has_error = True
                message = "Token amount error"
                break
            else:
                token_amount += 500
                continue
        except httpx.ReadTimeout as e:
            timeout += 30.0
            if timeout >= 120.0:
                has_error = True
                message = "Read timeout issue"
                break
            else:
                attempts += 1
                continue
        except Exception as e:
            has_error = True
            message = f"Unknown exception: {e}"
            break
    if has_error:
        send_to_error_channel(message, user.email, "get google summary (platform)")
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=message)
    return Response(
        data={
            "message": message,
            "results": results,
            "article": text,
        }
    )


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def read_column_names(request):
    file_obj = request.FILES.get("file")
    if file_obj:
        file_name = file_obj.name
        print(file_name)
        try:
            if file_name.endswith(".xlsx"):
                workbook = load_workbook(file_obj, data_only=True)
                res_data = []
                for sheet in workbook.sheetnames:
                    current_sheet = workbook[sheet]
                    column_names = [
                        cell.value
                        for cell in current_sheet[1]
                        if cell.value not in [None, False, True]
                    ]
                    res_data.append({"name": sheet, "columns": column_names})
            elif file_name.endswith(".xls"):
                workbook = xlrd.open_workbook(file_contents=file_obj.read())
                res_data = []
                for sheet in workbook.sheet_names():
                    current_sheet = workbook.sheet_by_name(sheet)
                    column_names = [
                        current_sheet.cell_value(0, col_idx)
                        for col_idx in range(current_sheet.ncols)
                    ]
                    res_data.append({"name": sheet, "columns": column_names})
            elif file_name.endswith(".csv"):
                csv_file = io.TextIOWrapper(file_obj.file, encoding="utf-8")
                reader = csv.reader(csv_file)
                column_names = next(reader)
                column_names = [col for col in column_names if col not in [None, "", False, True]]
                res_data = [{"name": file_name, "columns": column_names}]
            else:
                return Response(
                    {"error": "Unsupported file type"}, status=status.HTTP_400_BAD_REQUEST
                )
            return Response({"sheets": res_data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def process_excel_file(request):
    file_obj = request.FILES.get("file")
    sheet_name = request.data.get("sheet")
    labels = json.loads(request.data.get("labels"))
    tags = request.data.get("tag")
    if tags:
        tags = [tags]
    else:
        tags = None
    if file_obj:
        index_values = {}
        journalist_values = {}
        if file_obj.name.endswith(".xlsx"):
            workbook = load_workbook(file_obj, data_only=True)
            sheet = workbook[sheet_name]
            for key in labels.keys():
                label = labels[key]
                for cell in sheet[1]:
                    if cell.value == label:
                        index_values[cell.column] = key
            for idx in index_values.keys():
                row_values = []
                for row in sheet.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
                    value = row[0].strip() if isinstance(row[0], str) else row[0]
                    row_values.append(value)
                journalist_values[index_values[idx]] = row_values
        else:
            workbook = xlrd.open_workbook(file_contents=file_obj.read())
            sheet = workbook.sheet_by_name(sheet_name)
            for key in labels.keys():
                label = labels[key]
                for col_idx in range(sheet.ncols):
                    if sheet.cell_value(0, col_idx) == label:
                        index_values[col_idx] = key
            for col_idx in index_values.keys():
                row_values = []
                for row_idx in range(1, sheet.nrows):
                    value = sheet.cell_value(row_idx, col_idx)
                    value = value.strip() if isinstance(value, str) else value
                    row_values.append(value)
                journalist_values[index_values[col_idx]] = row_values
        filtered_emails = [email for email in journalist_values.get("email", []) if email]
        result = TaskResults.objects.create(
            function_name="emit_process_contacts_excel", user_id=str(request.user.id)
        )
        task = emit_process_contacts_excel(
            str(request.user.id), journalist_values, str(result.id), tags
        )
        result.task_id_str = str(task.id)
        result.task = task
        result.save()
        return Response(
            status=status.HTTP_200_OK,
            data={"task_id": str(result.id), "num_processing": len(filtered_emails)},
        )
    else:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def process_csv_file(request):
    file_obj = request.FILES.get("file")
    labels = json.loads(request.data.get("labels"))
    tags = request.data.get("tag", [])
    if tags:
        tags = [tags]
    if file_obj:
        index_values = {}
        journalist_values = {}
        csv_reader = csv.reader(file_obj.read().decode("utf-8").splitlines())
        header = next(csv_reader)
        for idx, cell in enumerate(header):
            if cell in labels.values():
                key = next(key for key, value in labels.items() if value == cell)
                index_values[idx] = key
        for row in csv_reader:
            for idx, key in index_values.items():
                if key not in journalist_values:
                    journalist_values[key] = []
                journalist_values[key].append(row[idx] if idx < len(row) else None)
        result = TaskResults.objects.create(
            function_name="emit_process_contacts_excel", user_id=str(request.user.id)
        )
        task = emit_process_contacts_excel(journalist_values, str(result.id), tags)
        result.task_id_str = str(task.id)
        result.task = task
        result.save()
        return Response(status=status.HTTP_204_NO_CONTENT)
    else:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)


# REDIRECTS


def redirect_from_email(request):
    verifier = request.GET.get("oauth_verifier", False)
    token = request.GET.get("oauth_token", False)
    q = urlencode({"state": "EMAIL", "oauth_verifier": verifier, "code": "code", "token": token})
    if not verifier:
        err = {"error": "there was an error"}
        err = urlencode(err)
        return redirect(f"{comms_consts.TWITTER_FRONTEND_REDIRECT}?{err}")
    return redirect(f"{comms_consts.TWITTER_FRONTEND_REDIRECT}?{q}")


def redirect_from_twitter(request):
    verifier = request.GET.get("oauth_verifier", False)
    token = request.GET.get("oauth_token", False)
    q = urlencode({"state": "TWITTER", "oauth_verifier": verifier, "code": "code", "token": token})
    if not verifier:
        err = {"error": "there was an error"}
        err = urlencode(err)
        return redirect(f"{comms_consts.TWITTER_FRONTEND_REDIRECT}?{err}")
    return redirect(f"{comms_consts.TWITTER_FRONTEND_REDIRECT}?{q}")


def redirect_from_instagram(request):
    code = request.GET.get("code", False)
    q = urlencode({"state": "INSTAGRAM", "code": code})
    if not code:
        err = {"error": "there was an error"}
        err = urlencode(err)
        return redirect(f"{comms_consts.TWITTER_FRONTEND_REDIRECT}?{err}")
    return redirect(f"{comms_consts.TWITTER_FRONTEND_REDIRECT}?{q}")


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_traffic_data(request):
    urls = request.data.get("urls")
    traffic_data = get_url_traffic_data(urls)
    emit_process_website_domain(urls, request.user.organization.name)
    if "error" in traffic_data.keys():
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=traffic_data)
    return Response(status=status.HTTP_200_OK, data=traffic_data)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_article_url_data(request):
    urls = request.data.get("urls")
    clip_data = get_article_data(urls)

    return Response(status=status.HTTP_200_OK, data=clip_data)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_social_url_data(request):
    urls = request.data.get("urls")
    clip_data = get_social_data(urls)
    return Response(status=status.HTTP_200_OK, data=clip_data)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_trending_articles(request):
    search = request.data.get("search")
    countries = request.data.get("countries")
    has_error = False
    attempts = 1
    token_amount = 1000
    timeout = 60.0

    while True:
        try:
            url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
            prompt = comms_consts.GENERATE_TREND_BOOLEAN(search)
            body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                "ed@mymanagr.com",
                prompt,
                "You are a VP of Communications",
                token_amount=token_amount,
                top_p=0.1,
                response_format={"type": "json_object"},
            )
            with Variable_Client(timeout) as client:
                r = client.post(
                    url,
                    data=json.dumps(body),
                    headers=core_consts.OPEN_AI_HEADERS,
                )
            res = open_ai_exceptions._handle_response(r)
            message = json.loads(res.get("choices")[0].get("message").get("content"))

            break
        except open_ai_exceptions.StopReasonLength:
            logger.exception(
                f"Retrying again due to token amount, amount currently at: {token_amount}"
            )
            if token_amount >= 2000:
                has_error = True

                message = "Token amount error"
                break
            else:
                token_amount += 500
                continue
        except httpx.ReadTimeout as e:
            print(e)
            timeout += 30.0
            if timeout >= 120.0:
                has_error = True
                message = "Read timeout issue"
                logger.exception(f"Read timeout from Open AI {e}")
                break
            else:
                attempts += 1
                continue
        except Exception as e:
            has_error = True
            message = f"Unknown exception: {e}"
            logger.exception(e)
            break
    if has_error:
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})

    topics = message["keywords"]

    articles = get_trend_articles(topics, countries)
    if "articles" in articles.keys():
        return Response(status=status.HTTP_200_OK, data={"articles": articles, "string": topics})
    else:
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=articles)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_clip_report_summary(request):
    elma = core_consts.ELMA
    user = request.user
    clips = request.data.get("clips")
    brand = request.data.get("brand")
    token_amount = 1000
    timeout = 60.0
    has_error = False
    while True:
        try:
            url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
            prompt = comms_consts.REPORT_SUMMARY(elma, brand, clips)
            body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                user.email,
                prompt,
                "You are a VP of Communications",
                token_amount=token_amount,
                top_p=0.1,
            )
            with Variable_Client(timeout) as client:
                r = client.post(
                    url,
                    data=json.dumps(body),
                    headers=core_consts.OPEN_AI_HEADERS,
                )
            res = open_ai_exceptions._handle_response(r)

            message = res.get("choices")[0].get("message").get("content").replace("**", "*")
            user.add_meta_data("generate_report")
            break
        except open_ai_exceptions.StopReasonLength:
            if token_amount >= 2000:
                has_error = True
                message = "Token amount error"
                break
            else:
                token_amount += 500
                continue
        except httpx.ReadTimeout as e:
            timeout += 30.0
            if timeout >= 120.0:
                has_error = True
                message = "Read timeout issue"
                break
            else:
                attempts += 1
                continue
        except Exception as e:
            has_error = True
            message = f"Unknown exception: {e}"
            break
    if has_error:
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
    return Response(status=status.HTTP_200_OK, data={"summary": message})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def rewrite_report(request):
    user = request.user
    content = request.data.get("content")
    instructions = request.data.get("instructions")
    clips = request.data.get("clips")
    has_error = False
    attempts = 1
    token_amount = 1000
    timeout = 60.0

    while True:
        try:
            url = core_consts.OPEN_AI_CHAT_COMPLETIONS_URI
            prompt = comms_consts.REGENERATE_REPORT_SUMMARY(content, instructions, clips)
            body = core_consts.OPEN_AI_CHAT_COMPLETIONS_BODY(
                user.email,
                prompt,
                "You are a VP of Communications",
                token_amount=token_amount,
                top_p=0.1,
            )
            with Variable_Client(timeout) as client:
                r = client.post(
                    url,
                    data=json.dumps(body),
                    headers=core_consts.OPEN_AI_HEADERS,
                )
            res = open_ai_exceptions._handle_response(r)

            message = res.get("choices")[0].get("message").get("content")

            break
        except open_ai_exceptions.StopReasonLength:
            logger.exception(
                f"Retrying again due to token amount, amount currently at: {token_amount}"
            )
            if token_amount >= 2000:
                has_error = True

                message = "Token amount error"
                break
            else:
                token_amount += 500
                continue
        except httpx.ReadTimeout as e:
            print(e)
            timeout += 30.0
            if timeout >= 120.0:
                has_error = True
                message = "Read timeout issue"
                logger.exception(f"Read timeout from Open AI {e}")
                break
            else:
                attempts += 1
                continue
        except Exception as e:
            has_error = True
            message = f"Unknown exception: {e}"
            logger.exception(e)
            break
    if has_error:
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data={"error": message})
    return Response(status=status.HTTP_200_OK, data={"summary": message})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_social_media_data(request):
    social_switcher = {
        "youtube": get_youtube_data,
        "twitter": get_tweet_data,
        "bluesky": get_bluesky_data,
    }
    user = request.user
    return_data = {}
    social_data_list = []
    errors = []
    params = request.data.get("params")
    query = params.get("query")
    project = params.get("project", None)
    converted_search = convert_social_search(query, user.email, project)
    max = 0
    social_values = ["youtube", "bluesky"]
    if user.has_twitter_integration:
        max = 25
        social_values.append("twitter")
    else:
        max = 50
    date_from = datetime.now(timezone.utc) - timedelta(days=7)
    for value in social_values:
        data_func = social_switcher[value]
        social_data = data_func(converted_search, max=max, user=user, date_from=date_from)
        if "error" in social_data.keys():
            errors.append(social_data["error"])
        else:
            if value == "twitter":
                return_data["includes"] = social_data["includes"]
            social_data_list.extend(social_data["data"])
    sorted_social_data = merge_sort_dates(social_data_list, "created_at")
    return_data["data"] = sorted_social_data
    if errors:
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=return_data)
    return Response(status=status.HTTP_200_OK, data=return_data)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
@authentication_classes([ExpiringTokenAuthentication])
def get_youtube_stats(request):
    video_id = request.data.get("video_id")
    headers = {"Accept": "application/json"}
    params = comms_consts.YOUTUBE_VIDEO_PARAMS(video_id)

    try:
        with Variable_Client(30) as client:
            res = client.get(comms_consts.YOUTUBE_VIDEO_URI, params=params, headers=headers)
            if res.status_code == 200:
                res = res.json()
                videos = res["items"][0]["statistics"]
            else:
                res = res.json()
                videos = {"error": res["error"]["message"]}
    except Exception as e:
        print(e)
        videos = {"error": str(e)}
        Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=videos)
    return Response(status=status.HTTP_200_OK, data=videos)
